#!/Library/Frameworks/Python.framework/Versions/3.9/bin/python3
"""
Création d'un chiffrier d'horaire pour PowerAutomate.
"""
import getopt
import os
import datetime
import sys
from zipfile import BadZipFile
import openpyxl

def affiche_aide():
    """
        Affiche l'aide pour la commande.
    """

    print("")
    print("""
    horaire.py -i <fichier_modele> -o <fichier_sortie> -h

    -i : Le chiffrier Excel contenant le modèle d'horaire.
    -o : Le fichier dans lequel sera créé l'horaire.
    -h : L'aide de la commande.
    """)

def heure_en_string(date_cours, heure_cours):
    """
    Convertit une date et une heure en string, format requis par PowerAutomate
    :param date_cours: datetime.date
    :param heure_cours: datetime.time
    :return: string
    """
    return datetime.datetime.combine(date_cours, heure_cours).strftime("%Y-%m-%dT%H:%M:00")

def creer_horaire(fichier_modele, fichier_sortie):
    """
    Lecture d'un modèle d'horaire (Excel) et
    création de toutes les entrées d'un cours pour la session.

        Paramètres
        ----------
        fichier_modele : str
            Nom et chemin du chiffrier Excel contenant l'horaire à créer.
        fichier_sortie : str
            Nom et chemin du chiffrier Excel à créer.
    """

    c_horaire = openpyxl.Workbook()
    f_horaire = c_horaire[c_horaire.sheetnames[0]]
    f_horaire.title = "Horaire"

    ligne = 1

    f_horaire.cell(row=ligne, column=1).value = "Sujet"
    f_horaire.cell(row=ligne, column=2).value = "Date début"
    f_horaire.cell(row=ligne, column=3).value = "Date fin"
    f_horaire.cell(row=ligne, column=4).value = "Emplacement"

    modele = openpyxl.load_workbook(fichier_modele, data_only=True)

    f_calendrier = modele["Calendrier"]
    f_cours = modele["Cours"]

    midi = datetime.time(12, 0, 0)

    for ligne_calendrier in range(2, f_calendrier.max_row + 1):
        for l_cours in range(2, f_cours.max_row + 1):
            mode_jour = f_calendrier.cell(row=ligne_calendrier, column=3).value
            mode_horaire = "AM" if f_cours.cell(row=l_cours, column=3).value < midi else "PM"

            if (
                mode_jour in ("COMPLET", mode_horaire) and
                f_calendrier.cell(row=ligne_calendrier, column=2).value ==
                  f_cours.cell(row=l_cours, column=2).value
                ):
                ligne = ligne + 1
                f_horaire.cell(row=ligne, column=1).value = f_cours.cell(
                                                                row=l_cours,
                                                                column=1).value
                f_horaire.cell(row=ligne, column=2).value = heure_en_string(
                    f_calendrier.cell(row=ligne_calendrier, column=1).value,
                    f_cours.cell(row=l_cours, column=3).value)
                f_horaire.cell(row=ligne, column=3).value = heure_en_string(
                    f_calendrier.cell(row=ligne_calendrier, column=1).value,
                    f_cours.cell(row=l_cours, column=4).value)
                f_horaire.cell(row=ligne, column=4).value = f_cours.cell(
                                                                row=l_cours,
                                                                column=5).value

    # définir le style de la table
    style_table = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                        showRowStripes=True)
    # Créer la table et l'affecter à la feuille
    f_horaire.add_table(openpyxl.worksheet.table.Table(ref=f'A1:D{ligne}',
                                        displayName='Horaires',
                                        tableStyleInfo=style_table))
    c_horaire.save(filename=fichier_sortie)

def valider_parametres(fichier_modele):
    """
        Valide l'ensemble des paramètres reçus en ligne de commande.
        Vérifie que le chiffrier contient bien les critères nécessaires.

        Paramètres
        ----------
        fichier_modele : str
            Nom et chemin du chiffrier Excel contenant l'horaire à créer.
        fichier_sortie : str
            Nom et chemin du chiffrier Excel à créer.

        Retour
        ------
        True si tout est valide.
    """

    parametres_valides = True

    # Validation des paramètres
    if not os.path.isfile(fichier_modele):
        print(f"Le fichier d'entrée {fichier_modele} n'existe pas.")
        parametres_valides = False

    # Vérifier si le fichier d'entrée est un chiffrier Excel
    try:
        chiffrier = openpyxl.load_workbook(fichier_modele, data_only=True)

        # Vérifier si la feuille existe
        if "Calendrier" not in chiffrier:
            print("La feuille Calendrier n'existe pas.")
            parametres_valides = False
        if "Cours" not in chiffrier:
            print("La feuille Cours n'existe pas.")
            parametres_valides = False
    except BadZipFile:
        print(f"Le fichier d'entrée {fichier_modele} n'est pas un chiffrier Excel valide.")
        parametres_valides = False

    return parametres_valides

def main(argv):
    """
        Procédure principale
    """

    fichier_modele = ''
    fichier_sortie = ''

    currentdir = os.getcwd()

    try:
        opts, _ = getopt.getopt(argv,"hi:o:")
    except getopt.GetoptError:
        affiche_aide()
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            affiche_aide()
            sys.exit()
        elif opt == '-i':
            fichier_modele = os.path.join(currentdir, arg)
        elif opt == '-o':
            fichier_sortie = os.path.join(currentdir, arg)

    if valider_parametres(fichier_modele):
        print(f'Fichier d\'entrée est : "{fichier_modele}"')
        print(f'Fichier de sortie est : "{fichier_sortie}"')
        creer_horaire(fichier_modele, fichier_sortie)

if __name__ == "__main__":
    main(sys.argv[1:])
