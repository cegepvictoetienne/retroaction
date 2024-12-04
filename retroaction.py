#!/Library/Frameworks/Python.framework/Versions/3.10/bin/python3
"""
 Générateur de rétroaction pour les élèves.

 Crée un fichier PDF pour chaque élève avec son numéro de demande d'admission (DA)
 à partir d'un chiffrier Excel qui contient les critères de correction (un critère
 par ligne dans la colonne A) et chaque élève par colonne (à partir de la colonne B)

"""
import getopt
import os
import sys

from pathlib import Path

from zipfile import BadZipFile
from zipfile import ZipFile

import openpyxl # type: ignore
from fpdf import FPDF # type: ignore
from fpdf.enums import XPos, YPos # type: ignore

# Constantes

# Quel est le caractère qui remplace le X pour indiquer que le critère est atteint
CROCHET = chr(214)
CROCHET_POLICE = "Symbol"
CROCHET_TAILLE = 14

LIBELLE_DA = "DA"
LIBELLE_NOTES = "Notes"
LIBELLE_NOM = "Nom"
LIBELLE_PRENOM = "Prénom"
LIBELLE_COMMENTAIRES = "Commentaires"
LIBELLE_SELECTION = "Générer"

HAUTEUR_CELLULE = 0.3
LARGEUR_TITRE = 6
LARGEUR_VALEUR = 2
DOSSIER_SCRIPT = Path( __file__ ).parent.absolute()
CHEMIN_LOGO = os.path.join(DOSSIER_SCRIPT, "logo.png")
CHEMIN_POLICE_REGULIER = os.path.join(DOSSIER_SCRIPT, "SourceSansPro-Regular.ttf")
CHEMIN_POLICE_GRAS = os.path.join(DOSSIER_SCRIPT, "SourceSansPro-Bold.ttf")
NOM_POLICE = "SourceSansPro"


class Eleve:
    """
        Classe contenant les données de l'élève et de ses résultats
    """
    def __init__(self, numero_da="0", prenom="", nom="", note=0):
        """
            Initialiser les données de l'élève.

            Paramètres
            ----------
            numero_da : str
                Numéro du dossier d'admission de l'élève
            prenom : str
                Prénom de l'élève
            nom : str
                Nom de famille de l'élève
            note : float
                Note finale de l'élève
        """
        self.prenom = prenom
        self.nom = nom
        self.note = note
        self.denominateur = 1
        self.commentaires = ""
        self.numero_da = numero_da
        self.notes = []

    def ajout_note(self, titre, valeur):
        """
            Ajouter une note à l'élève.

            Paramètres
            ----------
            titre : str
                Titre du critère évalué
            valeur : str
                Valeur du critère évalué
        """
        self.notes.append((titre, valeur))

    def nom_pdf(self):
        """
            Renvoyer le nom du fichier PDF pour l'élève.
        """
        return self.numero_da + ".pdf"

    def echec(self):
        """
            Renvoyer True si l'élève a echoué.
        """
        return self.note / self.denominateur < 0.6

    def afficher_note(self):
        """
            Afficher la note de l'élève.
        """
        return f"{self.note} / {self.denominateur} ({self.note_sur_100()} %)"

    def note_sur_100(self):
        """
            Renvoyer la note de l'élève sur 100.
        """
        return round(self.note / self.denominateur * 100)


class FeuilleEvaluation(FPDF):
    """
    Générer la page d'évaluation
    """
    def __init__(self, titre):
        """
            Initialiser la page d'évaluation

            Paramètres
            ----------
            titre : str
                Titre de la page
        """
        self.titre = titre

        super().__init__(orientation='P', unit='in', format="Letter")
        self.add_font(NOM_POLICE, fname=CHEMIN_POLICE_REGULIER)
        self.add_font(family=NOM_POLICE, style='B', fname=CHEMIN_POLICE_GRAS)


    def changer_police(self):
        """
        Changer la police de la page
        """
        self.set_font(NOM_POLICE, size=12)

    def changer_police_crochet(self):
        """
        Changer la police du crochet
        """
        self.set_font(CROCHET_POLICE, '', CROCHET_TAILLE)

    def header(self):
        """
        Générer l'entête de page
        """
        # Ajouter le logo du Cégep de Victoriaville
        self.image(CHEMIN_LOGO, 0, 0, 2)

        self.set_font(NOM_POLICE, 'B', 16)
        # Déplacer le curseur à droite
        self.cell(0.3)
        # Centrer le titre de la page

        self.cell(
            w=7,
            h=0.8,
            txt=self.titre,
            border=0,
            align='C',
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
            markdown=True
            )


    def footer(self):
        """
        Générer le pied de page
        """
        # Positionner le curseur à 1" du bas de page:
        self.set_y(-1)
        self.set_font(NOM_POLICE, 'B', 11)
        # Imprimer le numéro de page
        self.cell(
            w=7,
            h=HAUTEUR_CELLULE,
            txt=f"Page {self.page_no()} / {{nb}}",
            border=0,
            align='C',
            new_x=XPos.RIGHT,
            new_y=YPos.TOP,
            markdown=True
            )


    def ajouter_critere(self, titre_critere, valeur_critere):
        """
        Ajouter un critère à la page

        Paramètres
        ----------
        titre_critere : str
            Titre du critère
        valeur_critere : str
            Valeur du critère
        """

         # Générer le titre du critère, si pas de titre, pas de bordure
        bordure = 1
        if titre_critere == " ":
            # Si le champ est vide, ne pas afficher la bordure
            bordure = 0

        if valeur_critere is None:
            valeur_critere = " "
        else:
            valeur_critere = str(valeur_critere)

        if self.will_page_break(HAUTEUR_CELLULE*2):
            self.add_page()

        self.changer_police()

        old_position = {
            "x" : self.get_x(),
            "y" : self.get_y()
        }

        self.multi_cell(
            w=LARGEUR_TITRE,
            h=HAUTEUR_CELLULE,
            txt=titre_critere,
            border=bordure,
            align='L',
            new_x=XPos.RIGHT,
            new_y=YPos.NEXT,
            markdown=True
            )

        hauteur_valeur = self.get_y() - old_position["y"]

        # Ajuster la hauteur de la cellule de la valeur pour être identique
        # à la cellule du titre
        self.set_xy(old_position["x"] + LARGEUR_TITRE, old_position["y"])

        if valeur_critere in ("x", "X"):
            valeur_critere = CROCHET
            self.changer_police_crochet()

        self.multi_cell(
            w=LARGEUR_VALEUR,
            h=hauteur_valeur,
            txt=valeur_critere,
            border=bordure,
            align='C',
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
            )

    def ajouter_commentaire(self, titre, texte):
        """
        Ajouter un commentaire à la page

        Paramètres
        ----------
        titre : str
            Titre du commentaire
        texte : str
            Texte du commentaire
        """

        
        bordure = 1
        largeur_totale = LARGEUR_TITRE + LARGEUR_VALEUR

        if texte is None:
            valeur_critere = " "
        else:
            valeur_critere = str(texte)

        if self.will_page_break(HAUTEUR_CELLULE*2):
            self.add_page()

        self.changer_police()

        self.multi_cell(
            w=largeur_totale,
            h=HAUTEUR_CELLULE,
            txt=titre,
            border=bordure,
            align='L',
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
            markdown=True
            )

        self.multi_cell(
            w=largeur_totale,
            h=HAUTEUR_CELLULE,
            txt=valeur_critere,
            border=bordure,
            align='L',
            new_x=XPos.LMARGIN,
            new_y=YPos.NEXT,
            markdown=True
            )



def affiche_aide():
    """
        Affiche l'aide pour la commande.
    """

    print("")
    print(f"""
    retroaction.py -i <fichier_retro> -o <dossier_sortie> -s <nom_feuille> -d <denominateur> -p

    -i : Le chiffrier Excel contenant les rétroactions aux élèves. Chaque élément de la grille d'évaluation est en ligne et chaque élève est une colonne. Relatif au répertoire courant.
    -o : Le dossier dans lequel seront créés les pdf et l'archive zip.  Relatif au répertoire courant.
    -s : Le nom de la feuille contenant les rétroactions aux élèves.
    -d : Le dénominateur de la note de l'évaluation.
    -p : Exécution partielle avec une sélection en utilisant le critère {LIBELLE_SELECTION}
    """)

def traiter_eleve(dossier_sortie, eleve, titre_feuille):
    """
        Créer le PDF pour un élève.

        Paramètres
        ----------
        dossier_sortie : str
            Chemin sur disque du dossier qui recevra le PDF
        eleve : Eleve
            Objet représentant un élève
        titre_feuille : str
            Titre du document généré
        Retour
        ------
        nom_pdf : str
            Le nom du pdf créé
    """
    # Créer le PDF
    pdf = FeuilleEvaluation(titre_feuille)
    pdf.add_page()
    pdf.set_fill_color(r=255, g=255, b=255)

    # Imprimer les informations de l'élève
    pdf.ajouter_critere(LIBELLE_DA, eleve.numero_da)
    pdf.ajouter_critere(LIBELLE_NOM, eleve.nom)
    pdf.ajouter_critere(LIBELLE_PRENOM, eleve.prenom)
    pdf.ajouter_critere(LIBELLE_NOTES, eleve.afficher_note())
    pdf.ajouter_commentaire(LIBELLE_COMMENTAIRES, eleve.commentaires)

    # Traiter tous les critères de correction pour l'élève
    for ligne in eleve.notes:
        if '{texte}' in ligne[0]:
            titre = ligne[0].replace('{texte}', '')
            pdf.ajouter_commentaire(titre, ligne[1])
        else:
            pdf.ajouter_critere(ligne[0], ligne[1])

    # Écrire le PDF sur disque
    nom_pdf = os.path.join(dossier_sortie, eleve.nom_pdf())
    try:
        pdf.output(name=nom_pdf)
    except UnicodeEncodeError as erreur:
        print("Une erreur d'encodage du PDF lors de l'écriture du PDF suivant : ")
        print(nom_pdf)
        print(erreur)

    return nom_pdf


def trouver_lignes_criteres(feuille_a_traiter):
    """
    Paramètres
    ----------
    feuille_a_traiter : objet data_sheet
        La feuille Excel qui contient les rétroactions à traiter pour l'élève.

    Retour
    ------
    La liste des critères et leur ligne dans la feuille.
    """

    # Définir les critères à transférer
    criteres = {
        LIBELLE_NOM : 0,
        LIBELLE_PRENOM : 0,
        LIBELLE_DA : 0,
        LIBELLE_NOTES : 0,
        LIBELLE_SELECTION : 0,
        LIBELLE_COMMENTAIRES : 0,
        }

    # Trouver la ligne correspondante aux critères
    for cle, _ in criteres.items():
        for ligne in range(1, feuille_a_traiter.max_row + 1):
            if feuille_a_traiter.cell(column=1, row=ligne).value == cle:
                criteres[cle] = ligne
    return criteres


def sommaire_notes(eleves, dossier_sortie, denominateur, nom_feuille_a_traiter):
    """
        Écrire un chiffrier Excel avec la liste des DA et des notes

        Paramètres
        ----------
        feuille_a_traiter : objet data_sheet
            La feuille Excel qui contient les rétroactions à traiter pour l'élève.
        dossier_sortie : str
            Chemin sur disque du dossier qui recevra le PDF
        nom_feuille_a_traiter : str
            Le nom de la feuille Excel qui contient les rétroactions à traiter pour l'élève.
        denominateur : int
            Le dénominateur de la note totale
    """

    # Créer le chiffrier
    chiffrier = openpyxl.Workbook()
    feuille = chiffrier[chiffrier.sheetnames[0]]

    # Écrire les entêtes
    feuille.cell(row=1, column=1).value = LIBELLE_NOM
    feuille.cell(row=1, column=2).value = LIBELLE_PRENOM
    feuille.cell(row=1, column=3).value = LIBELLE_DA
    feuille.cell(row=1, column=4).value = f'Note sur {denominateur}'
    feuille.cell(row=1, column=5).value = 'Note sur 100'
    feuille.cell(row=1, column=6).value = 'Échec'

    ligne = 1
    for eleve in eleves:
        # Écrire les informations de l'élève
        ligne += 1
        feuille.cell(row=ligne, column=1).value = eleve.nom
        feuille.cell(row=ligne, column=2).value = eleve.prenom
        feuille.cell(row=ligne, column=3).value = eleve.numero_da
        feuille.cell(row=ligne, column=4).value = eleve.note
        feuille.cell(row=ligne, column=5).value = eleve.note_sur_100()
        feuille.cell(row=ligne, column=6).value = "Echec" if eleve.echec() else ""

    chiffrier.save(filename=f"{dossier_sortie}/{nom_feuille_a_traiter}.xlsx")


def generer_liste_eleves(fichier_retroaction,
                        nom_feuille_a_traiter,
                        denominateur,
                        traitement_partiel):
    """
        Désérialisation du chiffrier Excel en une liste d'objets de type Eleve

        Paramètres
        ----------
        fichier_retroaction : str
            Chemin du fichier Excel qui contient les rétroactions à traiter.
        nom_feuille_a_traiter : str
            Le nom de la feuille Excel qui contient les rétroactions à traiter pour l'élève.
        denominateur : int
            Le dénominateur de la note totale
        traitement_partiel : bool
            True si on doit traiter les rétroactions partiellement, False sinon.
    """
    # Ouvrir le chiffrier
    chiffrier = openpyxl.load_workbook(fichier_retroaction, data_only=True)
    feuille = chiffrier[nom_feuille_a_traiter]

    # Définir les critères à transférer
    criteres = trouver_lignes_criteres(feuille)

    # Créer la liste des élèves
    eleves = []

    # Traiter chaque étudiant
    for etudiant in range(2, feuille.max_column + 1):
        # Vérifier si le traitement partiel sélectionné est activé
        if (traitement_partiel and
            feuille.cell(column=etudiant, row=criteres[LIBELLE_SELECTION]).value != "X"):
            continue

        # Créer un objet élève
        eleve = Eleve()

        # Définir les valeurs
        eleve.nom = feuille.cell(column=etudiant, row=criteres[LIBELLE_NOM]).value
        eleve.prenom = feuille.cell(column=etudiant, row=criteres[LIBELLE_PRENOM]).value
        eleve.numero_da = str(feuille.cell(column=etudiant, row=criteres[LIBELLE_DA]).value)
        eleve.note = int(feuille.cell(column=etudiant, row=criteres[LIBELLE_NOTES]).value)
        eleve.commentaires = feuille.cell(column=etudiant, row=criteres[LIBELLE_COMMENTAIRES]).value
        eleve.denominateur = denominateur

        for element in range(criteres[LIBELLE_PRENOM] + 1, feuille.max_row + 1):

            titre_critere = feuille.cell(column=1, row=element).value
            if titre_critere is None:
                titre_critere = " "

            valeur_critere = feuille.cell(column=etudiant, row=element).value
            if valeur_critere is None:
                valeur_critere = " "

            eleve.ajout_note(titre_critere, valeur_critere)

        # Ajouter l'élève à la liste
        eleves.append(eleve)

    chiffrier.close()
    # Retourner la liste des élèves
    return eleves


def traiter_eleves(eleves, dossier_sortie, titre_feuille):
    """
    Traiter tous les élèves de la liste

    Paramètres
    ----------
    eleves : list
        La liste des élèves à traiter

    dossier_sortie : str
        Chemin du dossier qui recevra les fichiers PDF

    titre_feuille : str
        Le titre de la feuille Excel qui contient les rétroactions à traiter pour l'élève.
    """
    # Créer le fichier ZIP
    nom_zip = os.path.join(dossier_sortie, "travaux.zip")
    with ZipFile(nom_zip, "w") as fichier_zip:
        # Traiter chaque étudiant
        print(f"Création des fiches de rétroaction pour {len(eleves)} élève(s)")
        for eleve in eleves:
            fichier_zip.write(
                traiter_eleve(dossier_sortie, eleve, titre_feuille),
                    eleve.nom_pdf()
                    )

        fichier_zip.close()


def valider_parametres(fichier_retroaction, dossier_sortie, nom_feuille_a_traiter, denominateur):
    """
        Valide l'ensemble des paramètres reçus en ligne de commande.
        Vérifie que le chiffrier contient bien les critères nécessaires.

        Paramètres
        ----------
        fichier_retroaction : str
            Nom et chemin du chiffrier Excel contenant les rétroactions
        dossier_sortie : str
            Chemin sur disque du dossier qui recevra le PDF
        nom_feuille_a_traiter : str
            Le nom de la feuille Excel qui contient les rétroactions à traiter pour l'élève.
        denominateur : int
            Le dénominateur de la note totale

        Retour
        ------
        True si tout est valide.
    """

    parametres_valides = True

    # Validation des paramètres
    if not os.path.isfile(fichier_retroaction):
        print(f"Le fichier d'entrée {fichier_retroaction} n'existe pas.")
        parametres_valides = False

    # Vérifier si le fichier d'entrée est un chiffrier Excel
    try:
        chiffrier = openpyxl.load_workbook(fichier_retroaction, data_only=True)

        # Vérifier si la feuille existe
        if nom_feuille_a_traiter not in chiffrier:
            print(f"La feuille {nom_feuille_a_traiter} n'existe pas.")
            parametres_valides = False
        else:

            # Valider si les critères de base sont présents
            criteres = trouver_lignes_criteres(chiffrier[nom_feuille_a_traiter])

            for cle, valeur in criteres.items():
                if valeur == 0:
                    print(f"Le critère {cle} n'existe pas dans le chiffrier.")
                    parametres_valides = False

        chiffrier.close()
    except BadZipFile:
        print(f"Le fichier d'entrée {fichier_retroaction} n'est pas un chiffrier Excel valide.")
        parametres_valides = False

    if not os.path.isdir(dossier_sortie):
        print(f"Le dossier de sortie {dossier_sortie} n'existe pas.")
        parametres_valides = False

    if denominateur < 1:
        print("Le dénominateur doit être plus grand que zéro.")
        parametres_valides = False

    return parametres_valides

def mode_interactif():
    """
        Mode interactif
    """
    # Get a list of all files in the current directory
    files_in_directory = os.listdir('.')

    # Filter only files with .xlsx or .xls extensions
    fichiers_excel_dossier_courant = [file for file in files_in_directory if file.endswith(('.xlsx', '.xls'))]

    print("Rétroaction à partir de quel fichier?" )
    print("")
    for index, fichier in enumerate(fichiers_excel_dossier_courant):
        print(f'{index} - {fichier}')

    choix_fichier = int(input("?"))

    fichier_choisi = fichiers_excel_dossier_courant[choix_fichier]
    print(f'Fichier choisi : {fichier_choisi}')

    chiffrier = openpyxl.load_workbook(fichier_choisi, data_only=True)

    print("Rétroaction à partir de quel feuille?" )
    print("")
    for index, feuille in enumerate(chiffrier.worksheets):
        print(f'{index} - {feuille}')

    choix_feuille = int(input("?"))

    feuille_choisie = chiffrier.worksheets[choix_feuille].title
    print(f'Feuille choisie : {feuille_choisie}')

    print("Rétroaction dans quel dossier?" )
    print("")

    dossier = input("?")

    print("Dénominateur?" )
    print("")

    denominateur = int(input("?"))

    eleves = generer_liste_eleves(fichier_choisi, feuille_choisie,
    denominateur, False)
    traiter_eleves(eleves, dossier, feuille_choisie)
    sommaire_notes(eleves, dossier, denominateur, feuille_choisie)



    
def main(argv):
    """
        Procédure principale
    """

    fichier_retroaction = ''
    dossier_sortie = ''
    nom_feuille_a_traiter = ''
    denominateur = 0
    traitement_partiel = False
    titre_feuille = ""

    currentdir = os.getcwd()

    try:
        opts, _ = getopt.getopt(argv,"phi:o:s:d:t:")
    except getopt.GetoptError:
        affiche_aide()
        sys.exit(2)
    if len(opts) == 0:
        mode_interactif()
        return
    for opt, arg in opts:
        if opt == '-h':
            affiche_aide()
            sys.exit()
        elif opt == '-i':
            fichier_retroaction = os.path.join(currentdir, arg)
        elif opt == '-o':
            dossier_sortie = os.path.join(currentdir, arg)
        elif opt == '-s':
            nom_feuille_a_traiter = arg
        elif opt == '-t':
            titre_feuille = arg
        elif opt == '-d':
            denominateur = int(arg)
        elif opt == '-p':
            traitement_partiel = True

    if valider_parametres(fichier_retroaction, dossier_sortie, nom_feuille_a_traiter, denominateur):
        print(f'Fichier d\'entrée est : "{fichier_retroaction}"')
        print(f'Dossier de sortie est : "{dossier_sortie}"')
        print(f'Nom de la feuille est "{nom_feuille_a_traiter}"')
        if traitement_partiel:
            print("Traitement partiel")
        else:
            print("Traitement complet")
        print(f'La note est sur : {denominateur}')
        eleves = generer_liste_eleves(fichier_retroaction, nom_feuille_a_traiter,
            denominateur, traitement_partiel)
        traiter_eleves(eleves, dossier_sortie, titre_feuille)
        sommaire_notes(eleves, dossier_sortie, denominateur, nom_feuille_a_traiter)


if __name__ == "__main__":
    main(sys.argv[1:])
